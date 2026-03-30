"""Batch action persistence service."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas
from .batch_service import normalize_batch_no


def _find_action_by_normalized_batch(
    db: Session,
    snapshot_month: str,
    normalized_batch_no: str,
) -> models.BatchAction | None:
    """Find one action row by month + normalized batch key."""
    exact = (
        db.query(models.BatchAction)
        .filter(
            models.BatchAction.snapshot_month == snapshot_month,
            models.BatchAction.batch_no == normalized_batch_no,
        )
        .first()
    )
    if exact is not None:
        return exact

    rows = (
        db.query(models.BatchAction)
        .filter(models.BatchAction.snapshot_month == snapshot_month)
        .order_by(models.BatchAction.updated_at.desc(), models.BatchAction.id.desc())
        .all()
    )
    for row in rows:
        if normalize_batch_no(row.batch_no) == normalized_batch_no:
            return row
    return None


def save_or_update_action(
    db: Session,
    payload: schemas.ActionSaveRequest,
    updated_by: str,
) -> models.BatchAction:
    """Create or update one batch action record."""
    batch_no = normalize_batch_no(payload.batch_no)
    if not batch_no:
        raise ValueError("批次编号不能为空")

    action = _find_action_by_normalized_batch(
        db=db,
        snapshot_month=payload.snapshot_month,
        normalized_batch_no=batch_no,
    )

    if action is None:
        action = models.BatchAction(
            snapshot_month=payload.snapshot_month,
            batch_no=batch_no,
        )
        db.add(action)
    elif action.batch_no != batch_no:
        action.batch_no = batch_no

    action.reason_note = payload.reason_note
    action.responsible_dept = payload.responsible_dept
    action.action_plan = payload.action_plan
    action.action_status = normalize_action_status(payload.action_status)
    action.remark = payload.remark
    action.claim_weight_tons = payload.claim_weight_tons
    action.claim_amount = payload.claim_amount
    action.claim_currency = payload.claim_currency
    action.expected_completion = payload.expected_completion
    action.updated_by = updated_by
    db.flush()
    db.commit()
    db.refresh(action)
    return action


def carry_forward_actions(db: Session, new_month: str) -> dict:
    """Carry forward latest historical action to the new month by batch number.

    Rules:
    - target scope: batches in current month SAP snapshot
    - source scope: latest action in historical months (snapshot_month < new_month)
    - matching key: normalized batch number
    - keep current month action if it already exists (fill-empty, never overwrite)
    """
    new_batches = {
        normalize_batch_no(row[0])
        for row in db.query(models.InventorySnapshot.batch_no)
        .filter(models.InventorySnapshot.snapshot_month == new_month)
        .all()
        if normalize_batch_no(row[0])
    }
    if not new_batches:
        return {
            "candidate_batches": 0,
            "carried": 0,
            "history_matched": 0,
            "skipped_existing": 0,
            "skipped_no_history": 0,
        }

    existing_actions = {
        normalize_batch_no(row[0])
        for row in db.query(models.BatchAction.batch_no)
        .filter(models.BatchAction.snapshot_month == new_month)
        .all()
        if normalize_batch_no(row[0])
    }
    pending_batches = new_batches - existing_actions
    if not pending_batches:
        return {
            "candidate_batches": len(new_batches),
            "carried": 0,
            "history_matched": 0,
            "skipped_existing": len(existing_actions & new_batches),
            "skipped_no_history": 0,
        }

    # Newest-first ordering makes first hit per normalized batch the effective "latest".
    history_actions = (
        db.query(models.BatchAction)
        .filter(models.BatchAction.snapshot_month < new_month)
        .order_by(
            models.BatchAction.snapshot_month.desc(),
            models.BatchAction.updated_at.desc(),
            models.BatchAction.id.desc(),
        )
        .all()
    )

    latest_by_batch: dict[str, models.BatchAction] = {}
    for act in history_actions:
        normalized_batch = normalize_batch_no(act.batch_no)
        if not normalized_batch or normalized_batch not in pending_batches:
            continue
        if normalized_batch in latest_by_batch:
            continue
        latest_by_batch[normalized_batch] = act
        if len(latest_by_batch) >= len(pending_batches):
            break

    carried = 0
    for batch_no, act in latest_by_batch.items():
        db.add(
            models.BatchAction(
                snapshot_month=new_month,
                batch_no=batch_no,
                reason_note=act.reason_note,
                responsible_dept=act.responsible_dept,
                action_plan=act.action_plan,
                action_status=normalize_action_status(act.action_status),
                remark=act.remark,
                claim_weight_tons=act.claim_weight_tons,
                claim_amount=act.claim_amount,
                claim_currency=act.claim_currency,
                expected_completion=act.expected_completion,
                updated_by="system:carry-forward",
            )
        )
        carried += 1

    db.flush()
    return {
        "candidate_batches": len(new_batches),
        "carried": carried,
        "history_matched": len(latest_by_batch),
        "skipped_existing": len(existing_actions & new_batches),
        "skipped_no_history": len(pending_batches) - len(latest_by_batch),
    }


_STATUS_KEYWORDS: list[tuple[str, str]] = [
    ("已完成", "已完成"),
    ("已关闭", "已完成"),
    ("进行中", "进行中"),
    ("讨论中", "进行中"),
    ("待定", "待定"),
    ("待处理", "待定"),
    ("待办", "待定"),
    ("completed", "已完成"),
    ("done", "已完成"),
    ("in progress", "进行中"),
    ("pending", "待定"),
]


def _map_status(raw: str | None) -> tuple[str, str | None]:
    """Map raw status text into standardized status."""
    if not raw or not str(raw).strip():
        return "待定", None
    text = str(raw).strip()
    lower_text = text.lower()
    for keyword, std_val in _STATUS_KEYWORDS:
        if keyword.lower() in lower_text:
            return std_val, None
    return "待定", f"原始状态: {text}"


def normalize_action_status(raw: str | None) -> str:
    """Normalize any status text into the 3-state canonical set."""
    mapped, _ = _map_status(raw)
    return mapped


def _parse_claim_amount(raw: Any) -> float | None:
    """Parse claim amount, return None when invalid or empty."""
    if raw is None:
        return None
    try:
        if pd.isna(raw):
            return None
    except (TypeError, ValueError):
        pass
    try:
        val = float(raw)
        return val if val != 0 else None
    except (TypeError, ValueError):
        return None


def _parse_claim_weight_tons(raw: Any) -> float | None:
    """Parse claim tons, return None when invalid or empty."""
    return _parse_claim_amount(raw)


def _parse_expected_date(raw: Any) -> tuple[date | None, str | None]:
    """Parse expected completion date.

    Returns:
    - (date, None): valid date
    - (None, note): unparseable non-empty text
    - (None, None): empty
    """
    if raw is None:
        return None, None
    try:
        if pd.isna(raw):
            return None, None
    except (TypeError, ValueError):
        pass

    # Excel serial date.
    if isinstance(raw, (int, float)):
        try:
            parsed = pd.to_datetime(raw, unit="D", origin="1899-12-30", errors="coerce")
            if pd.notna(parsed):
                return parsed.date(), None
        except Exception:
            pass

    try:
        parsed = pd.to_datetime(raw, errors="coerce")
        if pd.notna(parsed):
            return parsed.date(), None
    except Exception:
        pass

    text = str(raw).strip()
    if text and text.lower() not in {"nan", "none", "nat"}:
        return None, f"预计完成: {text}"
    return None, None


def _safe_str(val: Any) -> str | None:
    """Convert any cell value to clean string, NaN-like values become None."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None
    return s


def _merge_action_fill_blanks(target: models.BatchAction, source: models.BatchAction) -> None:
    """Merge source into target using fill-empty-only strategy."""
    if not target.reason_note and source.reason_note:
        target.reason_note = source.reason_note
    if not target.responsible_dept and source.responsible_dept:
        target.responsible_dept = source.responsible_dept
    if not target.action_plan and source.action_plan:
        target.action_plan = source.action_plan
    target_status = normalize_action_status(target.action_status)
    source_status = normalize_action_status(source.action_status)
    if not target.action_status or target_status == "待定":
        target.action_status = source_status
    else:
        target.action_status = target_status
    if not target.remark and source.remark:
        target.remark = source.remark
    if target.claim_weight_tons is None and source.claim_weight_tons is not None:
        target.claim_weight_tons = source.claim_weight_tons
    if target.claim_amount is None and source.claim_amount is not None:
        target.claim_amount = source.claim_amount
    if not target.claim_currency and source.claim_currency:
        target.claim_currency = source.claim_currency
    if target.expected_completion is None and source.expected_completion is not None:
        target.expected_completion = source.expected_completion


def _read_action_excel_rows(raw: bytes) -> list[dict[str, Any]]:
    """Read action import rows with openpyxl to preserve text cell values."""
    workbook = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=False)
        header_cells = next(row_iter, None)
        if not header_cells:
            return []

        headers: list[str] = []
        for cell in header_cells:
            header = _safe_str(cell.value)
            headers.append(header or "")

        rows: list[dict[str, Any]] = []
        for cells in row_iter:
            row_data: dict[str, Any] = {}
            has_value = False
            for idx, cell in enumerate(cells):
                if idx >= len(headers):
                    continue
                header = headers[idx]
                if not header:
                    continue
                row_data[header] = cell.value
                if cell.value is not None:
                    has_value = True
            if has_value:
                rows.append(row_data)
        return rows
    finally:
        workbook.close()


def import_actions_from_excel(
    db: Session,
    file: UploadFile,
    snapshot_month: str,
    uploaded_by: str,
) -> schemas.ActionImportResponse:
    """Import offline action records and match abnormal batches by batch number."""
    raw = file.file.read()
    rows = _read_action_excel_rows(raw)
    if not rows:
        return schemas.ActionImportResponse(matched=0, skipped=0, errors=0, error_details=[])

    df = pd.DataFrame(rows)
    if "批次编号" not in df.columns:
        raise ValueError("Excel 缺少“批次编号”列")

    abnormal_batches = {
        normalize_batch_no(row[0])
        for row in db.query(models.InventorySnapshot.batch_no)
        .filter(
            models.InventorySnapshot.snapshot_month == snapshot_month,
            models.InventorySnapshot.quality_flag == "N",
        )
        .all()
        if normalize_batch_no(row[0])
    }

    # Existing actions of target month: normalize keys and merge duplicates first.
    existing_actions: dict[str, models.BatchAction] = {}
    existing_rows = (
        db.query(models.BatchAction)
        .filter(models.BatchAction.snapshot_month == snapshot_month)
        .order_by(models.BatchAction.updated_at.desc(), models.BatchAction.id.desc())
        .all()
    )
    grouped_actions: dict[str, list[models.BatchAction]] = {}
    for act in existing_rows:
        key = normalize_batch_no(act.batch_no)
        if key:
            grouped_actions.setdefault(key, []).append(act)

    for key, acts in grouped_actions.items():
        primary = next((a for a in acts if a.batch_no == key), acts[0])
        for duplicate in acts:
            if duplicate is primary:
                continue
            _merge_action_fill_blanks(primary, duplicate)
            db.delete(duplicate)
        if primary.batch_no != key:
            primary.batch_no = key
        primary.action_status = normalize_action_status(primary.action_status)
        existing_actions[key] = primary

    stats = schemas.ActionImportResponse(matched=0, skipped=0, errors=0, error_details=[])

    for idx, row in df.iterrows():
        row_num = idx + 2  # Header is row 1.
        try:
            batch_no = normalize_batch_no(row.get("批次编号"))
            if not batch_no:
                continue

            if batch_no not in abnormal_batches:
                stats.skipped += 1
                continue

            dept = _safe_str(row.get("责任部门"))
            if dept:
                dept = dept.replace("\\", "/")
            plan = _safe_str(row.get("处理方案"))
            raw_status = _safe_str(row.get("处理状态"))
            mapped_status, status_note = _map_status(raw_status)
            claim_weight_tons = _parse_claim_weight_tons(
                row.get("索赔吨数") or row.get("索赔重量")
            )
            claim = _parse_claim_amount(row.get("索赔金额"))
            claim_cur = _safe_str(row.get("币种"))
            exp_date, date_note = _parse_expected_date(row.get("预计完成时间"))
            reason_note = _safe_str(row.get("线下呆滞原因描述") or row.get("线下原因说明"))
            remark_raw = _safe_str(row.get("备注"))

            remark_parts = []
            if remark_raw:
                remark_parts.append(remark_raw)
            if status_note:
                remark_parts.append(status_note)
            if date_note:
                remark_parts.append(date_note)
            remark = "; ".join(remark_parts) if remark_parts else None

            if batch_no in existing_actions:
                act = existing_actions[batch_no]
                if not act.responsible_dept and dept:
                    act.responsible_dept = dept
                if not act.action_plan and plan:
                    act.action_plan = plan
                if not act.action_status or normalize_action_status(act.action_status) == "待定":
                    act.action_status = mapped_status
                else:
                    act.action_status = normalize_action_status(act.action_status)
                if not act.reason_note and reason_note:
                    act.reason_note = reason_note
                if act.claim_weight_tons is None and claim_weight_tons is not None:
                    act.claim_weight_tons = claim_weight_tons
                if act.claim_amount is None and claim is not None:
                    act.claim_amount = claim
                if not act.claim_currency and claim_cur:
                    act.claim_currency = claim_cur
                if act.expected_completion is None and exp_date is not None:
                    act.expected_completion = exp_date
                if not act.remark and remark:
                    act.remark = remark
                act.updated_by = uploaded_by
            else:
                new_act = models.BatchAction(
                    snapshot_month=snapshot_month,
                    batch_no=batch_no,
                    responsible_dept=dept,
                    action_plan=plan,
                    action_status=mapped_status,
                    reason_note=reason_note,
                    claim_weight_tons=claim_weight_tons,
                    claim_amount=claim,
                    claim_currency=claim_cur,
                    expected_completion=exp_date,
                    remark=remark,
                    updated_by=uploaded_by,
                )
                db.add(new_act)
                existing_actions[batch_no] = new_act

            stats.matched += 1
        except Exception as exc:
            stats.errors += 1
            stats.error_details.append(f"第{row_num}行: {exc}")
            if len(stats.error_details) > 50:
                stats.error_details.append("更多错误已省略")
                break

    db.flush()
    db.commit()
    return stats
